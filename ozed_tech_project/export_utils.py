"""
Export utilities for CSV, Excel, and PDF generation
"""
import csv
from io import BytesIO
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from datetime import datetime


class CSVExporter:
    """Export data to CSV format"""

    @staticmethod
    def export_to_csv(filename, headers, rows):
        """
        Export data to CSV

        Args:
            filename: Name of the file to download
            headers: List of column headers
            rows: List of data rows
        """
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        writer = csv.writer(response)
        writer.writerow(headers)
        writer.writerows(rows)

        return response


class ExcelExporter:
    """Export data to Excel format"""

    @staticmethod
    def export_to_excel(filename, sheet_name, headers, rows):
        """
        Export data to Excel

        Args:
            filename: Name of the file to download
            sheet_name: Name of the worksheet
            headers: List of column headers
            rows: List of data rows
        """
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name

        # Style for headers
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Write data rows
        for row_num, row_data in enumerate(rows, 2):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.alignment = Alignment(horizontal="left", vertical="center")

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Save to BytesIO
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response


class PDFExporter:
    """Export data to PDF format"""

    @staticmethod
    def create_invoice(sales_order):
        """
        Generate PDF invoice for a sales order

        Args:
            sales_order: SalesOrder model instance
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=72, leftMargin=72,
                                topMargin=72, bottomMargin=18)

        # Container for PDF elements
        elements = []

        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#667EEA'),
            spaceAfter=30,
            alignment=TA_CENTER
        )

        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#667EEA'),
            spaceAfter=12,
        )

        # Title
        title = Paragraph("SALES INVOICE", title_style)
        elements.append(title)
        elements.append(Spacer(1, 12))

        # Company Info
        company_info = f"""
        <b>Ozed Tech</b><br/>
        Inventory & CRM System<br/>
        Invoice Date: {datetime.now().strftime('%B %d, %Y')}
        """
        elements.append(Paragraph(company_info, styles['Normal']))
        elements.append(Spacer(1, 20))

        # Order Information
        order_info_data = [
            ['Order Number:', sales_order.order_number],
            ['Order Date:', sales_order.order_date.strftime('%B %d, %Y')],
            ['Status:', sales_order.get_status_display()],
            ['Payment Status:', sales_order.get_payment_status_display()],
        ]

        if sales_order.expected_delivery_date:
            order_info_data.append(['Expected Delivery:',
                                   sales_order.expected_delivery_date.strftime('%B %d, %Y')])

        order_info_table = Table(order_info_data, colWidths=[2*inch, 4*inch])
        order_info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F0F0F0')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))

        elements.append(order_info_table)
        elements.append(Spacer(1, 20))

        # Customer Information
        elements.append(Paragraph("Bill To:", heading_style))

        # Build customer info with available fields
        customer_lines = [f"<b>{sales_order.customer.company_name}</b>"]

        if sales_order.customer.address:
            customer_lines.append(sales_order.customer.address)

        # Add city, state, postal code line if any exist
        location_parts = [
            sales_order.customer.city or '',
            sales_order.customer.state or '',
            sales_order.customer.postal_code or ''
        ]
        location_line = ' '.join([part for part in location_parts if part]).strip()
        if location_line:
            customer_lines.append(location_line)

        if sales_order.customer.country:
            customer_lines.append(sales_order.customer.country)

        # Add contact information if contact is specified
        if sales_order.contact:
            if sales_order.contact.phone:
                customer_lines.append(f"Phone: {sales_order.contact.phone}")
            if sales_order.contact.email:
                customer_lines.append(f"Email: {sales_order.contact.email}")
            customer_lines.append(f"Attn: {sales_order.contact.full_name}")

        # Add website if available
        if sales_order.customer.website:
            customer_lines.append(f"Website: {sales_order.customer.website}")

        customer_info = "<br/>".join(customer_lines)
        elements.append(Paragraph(customer_info, styles['Normal']))
        elements.append(Spacer(1, 20))

        # Items Table
        elements.append(Paragraph("Order Items:", heading_style))

        items_data = [['Item', 'SKU', 'Quantity', 'Unit Price', 'Discount', 'Total']]

        for item in sales_order.items.all():
            items_data.append([
                item.item.name,
                item.item.sku or 'N/A',
                str(item.quantity),
                f'${item.unit_price:.2f}',
                f'${item.discount:.2f}',
                f'${item.subtotal:.2f}'
            ])

        items_table = Table(items_data, colWidths=[2*inch, 1*inch, 0.8*inch, 1*inch, 0.8*inch, 1*inch])
        items_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667EEA')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ]))

        elements.append(items_table)
        elements.append(Spacer(1, 20))

        # Totals
        totals_data = [
            ['Subtotal:', f'${sales_order.subtotal:.2f}'],
            ['Discount:', f'-${sales_order.discount:.2f}'],
            ['Tax:', f'${sales_order.tax:.2f}'],
            ['Shipping:', f'${sales_order.shipping_cost:.2f}'],
            ['', ''],
            ['Total Amount:', f'${sales_order.total_amount:.2f}'],
        ]

        totals_table = Table(totals_data, colWidths=[4.5*inch, 2*inch])
        totals_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 14),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.HexColor('#667EEA')),
            ('LINEABOVE', (0, -1), (-1, -1), 2, colors.HexColor('#667EEA')),
            ('TOPPADDING', (0, -1), (-1, -1), 10),
        ]))

        elements.append(totals_table)
        elements.append(Spacer(1, 30))

        # Notes
        if sales_order.notes:
            elements.append(Paragraph("Notes:", heading_style))
            elements.append(Paragraph(sales_order.notes, styles['Normal']))
            elements.append(Spacer(1, 20))

        # Footer
        footer_text = """
        <para align=center>
        <b>Thank you for your business!</b><br/>
        For any questions regarding this invoice, please contact us.<br/>
        Generated by Ozed Tech Inventory & CRM System
        </para>
        """
        elements.append(Spacer(1, 30))
        elements.append(Paragraph(footer_text, styles['Normal']))

        # Build PDF
        doc.build(elements)

        # Get PDF from buffer
        pdf = buffer.getvalue()
        buffer.close()

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="invoice_{sales_order.order_number}.pdf"'
        response.write(pdf)

        return response

    @staticmethod
    def create_quote_pdf(quote):
        """
        Generate PDF for a quote

        Args:
            quote: Quote model instance
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=72, leftMargin=72,
                                topMargin=72, bottomMargin=18)

        # Container for PDF elements
        elements = []

        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#667EEA'),
            spaceAfter=30,
            alignment=TA_CENTER
        )

        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#667EEA'),
            spaceAfter=12,
        )

        # Title
        title = Paragraph(f"SALES QUOTE #{quote.quote_number}", title_style)
        elements.append(title)

        # Version badge if not version 1
        if quote.version > 1:
            version_text = f"<para align=center><b>Version {quote.version}</b></para>"
            elements.append(Paragraph(version_text, styles['Normal']))

        elements.append(Spacer(1, 12))

        # Company Info
        company_info = f"""
        <b>Ozed Tech</b><br/>
        Inventory & CRM System<br/>
        Quote Date: {quote.quote_date.strftime('%B %d, %Y')}
        """
        elements.append(Paragraph(company_info, styles['Normal']))
        elements.append(Spacer(1, 20))

        # Quote Information
        quote_info_data = [
            ['Quote Number:', quote.quote_number],
            ['Version:', str(quote.version)],
            ['Quote Date:', quote.quote_date.strftime('%B %d, %Y')],
            ['Expiration Date:', quote.expiration_date.strftime('%B %d, %Y')],
            ['Status:', quote.get_status_display()],
            ['Payment Terms:', quote.get_payment_terms_display()],
            ['Validity Period:', quote.validity_period],
        ]

        if quote.rfq:
            quote_info_data.insert(1, ['Reference RFQ:', quote.rfq.rfq_number])

        quote_info_table = Table(quote_info_data, colWidths=[2*inch, 4*inch])
        quote_info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F0F0F0')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))

        elements.append(quote_info_table)
        elements.append(Spacer(1, 20))

        # Customer Information
        elements.append(Paragraph("Quote For:", heading_style))

        # Build customer info with available fields
        customer_lines = [f"<b>{quote.customer.company_name}</b>"]

        if quote.customer.address:
            customer_lines.append(quote.customer.address)

        # Add city, state, postal code line if any exist
        location_parts = [
            quote.customer.city or '',
            quote.customer.state or '',
            quote.customer.postal_code or ''
        ]
        location_line = ' '.join([part for part in location_parts if part]).strip()
        if location_line:
            customer_lines.append(location_line)

        if quote.customer.country:
            customer_lines.append(quote.customer.country)

        # Add contact information if contact is specified
        if quote.contact:
            if quote.contact.phone:
                customer_lines.append(f"Phone: {quote.contact.phone}")
            if quote.contact.email:
                customer_lines.append(f"Email: {quote.contact.email}")
            customer_lines.append(f"Attn: {quote.contact.full_name}")

        # Add sales rep info
        if quote.sales_rep:
            customer_lines.append(f"<br/><b>Sales Representative:</b> {quote.sales_rep.username}")

        customer_info = "<br/>".join(customer_lines)
        elements.append(Paragraph(customer_info, styles['Normal']))
        elements.append(Spacer(1, 20))

        # Items Table
        elements.append(Paragraph("Quoted Items:", heading_style))

        items_data = [['Item', 'SKU', 'Quantity', 'Unit Price', 'Discount', 'Total']]

        for item in quote.items.all():
            items_data.append([
                item.item.name,
                item.item.sku or 'N/A',
                str(item.quantity),
                f'${item.unit_price:.2f}',
                f'${item.discount:.2f}',
                f'${item.subtotal:.2f}'
            ])

        items_table = Table(items_data, colWidths=[2*inch, 1*inch, 0.8*inch, 1*inch, 0.8*inch, 1*inch])
        items_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667EEA')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ]))

        elements.append(items_table)
        elements.append(Spacer(1, 20))

        # Totals
        totals_data = [
            ['Subtotal:', f'${quote.subtotal:.2f}'],
            ['Discount:', f'-${quote.discount:.2f}'],
            ['Tax:', f'${quote.tax:.2f}'],
            ['Shipping:', f'${quote.shipping_cost:.2f}'],
            ['', ''],
            ['Total Amount:', f'${quote.total_amount:.2f}'],
        ]

        totals_table = Table(totals_data, colWidths=[4.5*inch, 2*inch])
        totals_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 14),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.HexColor('#667EEA')),
            ('LINEABOVE', (0, -1), (-1, -1), 2, colors.HexColor('#667EEA')),
            ('TOPPADDING', (0, -1), (-1, -1), 10),
        ]))

        elements.append(totals_table)
        elements.append(Spacer(1, 30))

        # Delivery Terms
        if quote.delivery_terms:
            elements.append(Paragraph("Delivery Terms:", heading_style))
            elements.append(Paragraph(quote.delivery_terms, styles['Normal']))
            elements.append(Spacer(1, 15))

        # Notes
        if quote.notes:
            elements.append(Paragraph("Terms & Conditions:", heading_style))
            elements.append(Paragraph(quote.notes, styles['Normal']))
            elements.append(Spacer(1, 20))

        # Footer
        footer_text = f"""
        <para align=center>
        <b>This quote is valid until {quote.expiration_date.strftime('%B %d, %Y')}</b><br/>
        <br/>
        Please contact us if you have any questions or would like to discuss this quote.<br/>
        We look forward to serving your needs!<br/>
        <br/>
        Generated by Ozed Tech Inventory & CRM System
        </para>
        """
        elements.append(Spacer(1, 20))
        elements.append(Paragraph(footer_text, styles['Normal']))

        # Build PDF
        doc.build(elements)

        # Get PDF from buffer
        pdf = buffer.getvalue()
        buffer.close()

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="quote_{quote.quote_number}_v{quote.version}.pdf"'
        response.write(pdf)

        return response
